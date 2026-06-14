"""Tests for ingestion: formats, name preservation, normalization, and error paths."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook

from veritas.ingest import IngestError, ingest_file, normalize_column_names
from veritas.session import InvestigationSession

FIXTURES = Path(__file__).parent / "fixtures"


def test_csv_ingest_basics(session: InvestigationSession, trio_paths: dict[str, Path]) -> None:
    record = ingest_file(session, trio_paths["csv"], name="trio")
    assert record.dataset_id.startswith("ds_")
    assert record.table_name == record.dataset_id
    assert (record.name, record.source_format) == ("trio", "csv")
    assert (record.row_count, record.column_count) == (6, 5)
    assert [c.normalized_name for c in record.schema_record.columns] == [
        "id",
        "category",
        "amount",
        "day",
        "qty",
    ]
    assert [c.duckdb_type for c in record.schema_record.columns] == [
        "BIGINT",
        "VARCHAR",
        "DOUBLE",
        "DATE",
        "BIGINT",
    ]
    row = session.conn.execute(f'SELECT count(*) FROM "{record.table_name}"').fetchone()
    assert row is not None and row[0] == 6
    assert session.get_dataset(record.dataset_id) == record


def test_name_defaults_to_file_stem(
    session: InvestigationSession, trio_paths: dict[str, Path]
) -> None:
    assert ingest_file(session, trio_paths["csv"]).name == "trio"


def test_duplicate_columns_preserved_and_deduplicated(session: InvestigationSession) -> None:
    record = ingest_file(session, FIXTURES / "duplicate_cols.csv")
    schema = record.schema_record
    assert [c.original_name for c in schema.columns] == ["day", "revenue", "revenue"]
    assert [c.normalized_name for c in schema.columns] == ["day", "revenue", "revenue_2"]
    assert schema.normalized_for("revenue") == ["revenue", "revenue_2"]
    assert schema.original_for("revenue_2") == "revenue"
    row = session.conn.execute(
        f'SELECT revenue, revenue_2 FROM "{record.table_name}" ORDER BY day LIMIT 1'
    ).fetchone()
    assert row == (100, 200)


def test_normalize_column_names_rules() -> None:
    reserved = frozenset({"select"})
    assert normalize_column_names(["Revenue (USD)", "revenue usd", "select"], reserved) == [
        "revenue_usd",
        "revenue_usd_2",
        "select_",
    ]
    assert normalize_column_names(["123abc", "", "café"], reserved) == [
        "c_123abc",
        "col_1",
        "caf",
    ]
    # deterministic positional suffixes for triplicates
    assert normalize_column_names(["x", "x", "x"], reserved) == ["x", "x_2", "x_3"]


def test_empty_csv_header_only(session: InvestigationSession) -> None:
    record = ingest_file(session, FIXTURES / "empty.csv")
    assert (record.row_count, record.column_count) == (0, 2)
    assert [c.normalized_name for c in record.schema_record.columns] == ["id", "value"]


def test_missing_file_raises(session: InvestigationSession) -> None:
    with pytest.raises(IngestError, match="not found"):
        ingest_file(session, "nope/missing.csv")


def test_unsupported_extension_raises(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "data.json"
    path.write_text("{}")
    with pytest.raises(IngestError, match="unsupported file type"):
        ingest_file(session, path)


def test_zero_byte_csv_raises(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "zero.csv"
    path.write_text("")
    with pytest.raises(IngestError, match="header row is required"):
        ingest_file(session, path)


def test_ragged_csv_rejected(session: InvestigationSession, tmp_path: Path) -> None:
    # the sniffer would skip the real header and fabricate originals from data cells
    path = tmp_path / "ragged.csv"
    path.write_text("a,b\n1,2,3\n4,5,6\n")
    with pytest.raises(IngestError, match="skips the first"):
        ingest_file(session, path)


def test_column_named_like_rename_placeholder(
    session: InvestigationSession, tmp_path: Path
) -> None:
    # regression: a column literally named like a phase-1 placeholder must not collide
    path = tmp_path / "tricky.csv"
    path.write_text("a,_veritas_tmp_0\n1,2\n")
    record = ingest_file(session, path)
    assert [c.original_name for c in record.schema_record.columns] == ["a", "_veritas_tmp_0"]
    assert [c.normalized_name for c in record.schema_record.columns] == ["a", "veritas_tmp_0"]
    row = session.conn.execute(f'SELECT a, veritas_tmp_0 FROM "{record.table_name}"').fetchone()
    assert row == (1, 2)


def test_timestamptz_csv_normalized_to_utc(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "tz.csv"
    path.write_text("ts,v\n2024-01-01 00:00:00+00,1\n2024-01-02 05:30:00+02,2\n")
    record = ingest_file(session, path)
    ts_col = record.schema_record.columns[0]
    assert ts_col.duckdb_type == "TIMESTAMP"  # D-014: stored as naive UTC
    row = session.conn.execute(f'SELECT min(ts), max(ts) FROM "{record.table_name}"').fetchone()
    assert row is not None
    assert row[0].isoformat() == "2024-01-01T00:00:00"
    assert row[1].isoformat() == "2024-01-02T03:30:00"  # +02 wall time → UTC


def test_timestamptz_parquet_normalized_to_utc(
    session: InvestigationSession, tmp_path: Path
) -> None:
    path = tmp_path / "tz.parquet"
    scratch = duckdb.connect()
    escaped = path.as_posix().replace("'", "''")
    scratch.execute(
        "COPY (SELECT TIMESTAMPTZ '2024-01-01 10:00:00+02' AS ts, 1 AS v) "
        f"TO '{escaped}' (FORMAT PARQUET)"
    )
    scratch.close()
    record = ingest_file(session, path)
    assert record.schema_record.columns[0].duckdb_type == "TIMESTAMP"
    row = session.conn.execute(f'SELECT ts FROM "{record.table_name}"').fetchone()
    assert row is not None and row[0].isoformat() == "2024-01-01T08:00:00"


def test_xlsx_date_with_null_becomes_date(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "dates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["day"])
    sheet.append([date(2024, 1, 1)])
    sheet.append([None])
    sheet.append([date(2024, 1, 3)])
    workbook.save(path)
    record = ingest_file(session, path)
    assert record.schema_record.columns[0].duckdb_type == "DATE"
    row = session.conn.execute(
        f'SELECT count(*), count(day), min(day) FROM "{record.table_name}"'
    ).fetchone()
    assert row is not None
    assert (row[0], row[1]) == (3, 2)
    assert row[2].isoformat() == "2024-01-01"


def test_corrupt_xlsx_raises(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "broken.xlsx"
    path.write_text("this is not a zip archive")
    with pytest.raises(IngestError, match="could not read Excel"):
        ingest_file(session, path)


def test_corrupt_parquet_raises(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "broken.parquet"
    path.write_text("this is not parquet")
    with pytest.raises(IngestError, match="could not ingest"):
        ingest_file(session, path)


def test_fully_empty_xlsx_sheet_raises(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "blank.xlsx"
    Workbook().save(path)
    with pytest.raises(IngestError, match="no header row"):
        ingest_file(session, path)


def test_parquet_ingest_matches_csv_schema(
    session: InvestigationSession, trio_paths: dict[str, Path]
) -> None:
    csv_record = ingest_file(session, trio_paths["csv"])
    parquet_record = ingest_file(session, trio_paths["parquet"])
    assert parquet_record.source_format == "parquet"
    assert parquet_record.row_count == csv_record.row_count
    assert [(c.normalized_name, c.duckdb_type) for c in parquet_record.schema_record.columns] == [
        (c.normalized_name, c.duckdb_type) for c in csv_record.schema_record.columns
    ]


def test_xlsx_ingest_matches_csv_schema(
    session: InvestigationSession, trio_paths: dict[str, Path]
) -> None:
    csv_record = ingest_file(session, trio_paths["csv"])
    xlsx_record = ingest_file(session, trio_paths["xlsx"])
    assert xlsx_record.source_format == "xlsx"
    assert xlsx_record.row_count == csv_record.row_count
    assert [(c.normalized_name, c.duckdb_type) for c in xlsx_record.schema_record.columns] == [
        (c.normalized_name, c.duckdb_type) for c in csv_record.schema_record.columns
    ]


def test_xlsx_uses_no_duckdb_extensions(
    session: InvestigationSession, trio_paths: dict[str, Path]
) -> None:
    ingest_file(session, trio_paths["xlsx"])
    loaded = {
        row[0]
        for row in session.conn.execute(
            "SELECT extension_name FROM duckdb_extensions() WHERE loaded"
        ).fetchall()
    }
    assert not loaded & {"excel", "spatial"}


def test_xlsx_duplicate_headers(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "dup.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["x", "x"])
    sheet.append([1, 2])
    workbook.save(path)
    record = ingest_file(session, path)
    schema = record.schema_record
    assert [c.original_name for c in schema.columns] == ["x", "x"]
    assert [c.normalized_name for c in schema.columns] == ["x", "x_2"]
    row = session.conn.execute(f'SELECT x, x_2 FROM "{record.table_name}"').fetchone()
    assert row == (1, 2)


def test_xlsx_header_only_sheet(session: InvestigationSession, tmp_path: Path) -> None:
    path = tmp_path / "header_only.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["a", "b"])
    workbook.save(path)
    record = ingest_file(session, path)
    assert (record.row_count, record.column_count) == (0, 2)
